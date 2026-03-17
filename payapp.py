import io
import re
import hmac
import hashlib
from datetime import datetime
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="CT Data Transformer",
    page_icon="🔄",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #f7f5f2; }
.block-container { padding: 2.5rem 2rem 4rem 2rem; max-width: 1100px; }
#MainMenu, footer, header { visibility: hidden; }
.stDeployButton { display: none; }
.login-box { max-width: 420px; margin: 60px auto 0; padding: 44px 48px; background: #fff; border: 1px solid #e5e1db; border-radius: 10px; border-top: 3px solid #d4521a; box-shadow: 0 4px 24px rgba(0,0,0,0.07); }
.login-icon { font-size: 30px; margin-bottom: 8px; }
.login-title { font-size: 22px; font-weight: 700; color: #1a1714; margin-bottom: 4px; }
.login-sub { font-size: 13px; color: #8a7f75; margin-bottom: 28px; }
.admin-badge { display:inline-block; background:#d4521a; color:#fff; font-size:10px; font-weight:700; letter-spacing:.06em; text-transform:uppercase; padding:2px 8px; border-radius:3px; margin-left:6px; }
.hero { background: #1a1714; border-radius: 8px; padding: 36px 44px 32px; margin-bottom: 28px; position: relative; overflow: hidden; }
.hero::before { content: ''; position: absolute; top: -60px; right: -60px; width: 280px; height: 280px; background: radial-gradient(circle, rgba(212,82,26,0.25) 0%, transparent 70%); pointer-events: none; }
.hero-label { font-size: 11px; font-weight: 600; letter-spacing: .12em; text-transform: uppercase; color: #d4521a; display: flex; align-items: center; gap: 8px; margin-bottom: 12px; }
.hero-label::before { content: ''; width: 20px; height: 2px; background: #d4521a; }
.hero-title { font-size: 34px; font-weight: 300; color: #f5f0eb; letter-spacing: -.02em; margin-bottom: 6px; }
.hero-title strong { font-weight: 700; }
.hero-sub { font-size: 13px; color: #8a7f75; line-height: 1.6; max-width: 520px; }
.hero-user { position: absolute; top: 24px; right: 28px; font-size: 12px; color: #8a7f75; }
.hero-user strong { color: #d4521a; }
.metric-row { display: flex; gap: 12px; margin-bottom: 24px; }
.metric-card { flex: 1; background: #fff; border: 1px solid #e5e1db; border-radius: 6px; padding: 18px 20px; border-left: 3px solid #d4521a; }
.metric-val { font-size: 26px; font-weight: 700; font-family: 'DM Mono', monospace; color: #d4521a; line-height: 1; }
.metric-label { font-size: 10px; font-weight: 600; letter-spacing: .08em; text-transform: uppercase; color: #8a7f75; margin-top: 4px; }
.section-head { font-size: 10px; font-weight: 600; letter-spacing: .1em; text-transform: uppercase; color: #8a7f75; display: flex; align-items: center; gap: 8px; margin-bottom: 14px; padding-bottom: 10px; border-bottom: 1px solid #e5e1db; }
.section-head span { background: #d4521a; color: #fff; font-size: 9px; padding: 2px 7px; border-radius: 2px; }
[data-testid="stFileUploadDropzone"] { background: #fff !important; border: 2px dashed #d4c8bc !important; border-radius: 6px !important; }
[data-testid="stFileUploadDropzone"]:hover { border-color: #d4521a !important; }
[data-testid="stSelectbox"] > div > div { background: #fff !important; border: 1px solid #e5e1db !important; border-radius: 4px !important; font-size: 13px !important; }
[data-testid="stTextInput"] input { border: 1px solid #e5e1db !important; border-radius: 4px !important; font-size: 14px !important; }
[data-testid="stTextInput"] input:focus { border-color: #d4521a !important; box-shadow: none !important; }
.stButton > button, .stDownloadButton > button { border-radius: 4px !important; font-family: 'DM Sans', sans-serif !important; font-weight: 600 !important; font-size: 13px !important; transition: all .2s !important; }
.stButton > button[kind="primary"], .stDownloadButton > button { background: #d4521a !important; border: none !important; color: #fff !important; }
.stButton > button[kind="primary"]:hover, .stDownloadButton > button:hover { background: #c4420a !important; transform: translateY(-1px) !important; box-shadow: 0 4px 16px rgba(212,82,26,0.3) !important; }
.stSuccess { background: rgba(42,122,75,.07) !important; border-color: rgba(42,122,75,.25) !important; border-radius: 4px !important; }
.stInfo { background: rgba(212,82,26,.06) !important; border-color: rgba(212,82,26,.2) !important; border-radius: 4px !important; }
.stError, .stWarning { border-radius: 4px !important; }
[data-testid="stDataFrame"] { border: 1px solid #e5e1db !important; border-radius: 6px !important; overflow: hidden; }
hr { border-color: #e5e1db !important; }
[data-testid="stExpander"] { background: #fff !important; border: 1px solid #e5e1db !important; border-radius: 6px !important; }
</style>
""", unsafe_allow_html=True)

# ── AUTH ──────────────────────────────────────────────────────────────────────

ADMIN_EMAIL = "vivek.naiwal@cars24.com"

def get_users():
    try:
        return {k.strip().lower(): v for k, v in dict(st.secrets["users"]).items()}
    except Exception:
        return {"vivek.naiwal@cars24.com": "Vivek@007"}

def check_login(email, password):
    users = get_users()
    email = email.strip().lower()
    if email not in users:
        return False
    return hmac.compare_digest(
        hashlib.sha256(password.encode()).hexdigest(),
        hashlib.sha256(str(users[email]).encode()).hexdigest()
    )

def show_login():
    _, center, _ = st.columns([1, 1.4, 1])
    with center:
        st.markdown("""
        <div class="login-box">
            <div class="login-icon">🔐</div>
            <div class="login-title">Payroll Portal</div>
            <div class="login-sub">CT Data Transformer — Restricted Access<br>Cars24 Internal Use Only</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        email    = st.text_input("Email", placeholder="you@cars24.com", key="li_email")
        password = st.text_input("Password", type="password", placeholder="Enter password", key="li_pass")
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Sign In →", type="primary", use_container_width=True):
            if not email or not password:
                st.error("Please enter both email and password.")
            elif check_login(email, password):
                st.session_state["auth"]       = True
                st.session_state["user_email"] = email.strip().lower()
                st.session_state["is_admin"]   = (email.strip().lower() == ADMIN_EMAIL.lower())
                st.rerun()
            else:
                st.error("❌ Wrong email or password. Contact vivek.naiwal@cars24.com")
        st.markdown("<div style='text-align:center;font-size:11px;color:#b0a89f;margin-top:16px;'>Authorised personnel only</div>", unsafe_allow_html=True)

if not st.session_state.get("auth"):
    show_login()
    st.stop()

user_email = st.session_state["user_email"]
is_admin   = st.session_state["is_admin"]

# ── COLUMN CANDIDATES ─────────────────────────────────────────────────────────

COLUMN_CANDIDATES = {
    "emp_id":       ["emp id", "empid", "employee id", "ecode", "emp_id", "employee_code"],
    "emp_name":     ["employee name", "name", "emp name", "employee"],
    "when_change":  ["when was the change", "when_was_the_change", "effective date", "when changed", "change date", "whenchange"],
    "total_ctc":    ["total ctc", "total ct", "ctc", "total_ctc", "total_ct", "total"],
    "fixed_ctc":    ["fixed ctc", "fixed_ctc", "fixed ct", "fixed salary", "fixed"],
    "variable_pay": ["total variable pay", "variable pay", "total_variable_pay", "variable", "var pay"],
    "created_date": ["created date", "created_date", "created on", "created", "created_at", "createdat"]
}

# ── HELPERS ───────────────────────────────────────────────────────────────────

def find_column(columns, candidates):
    cols_map = {c.lower().strip(): c for c in columns}
    for cand in candidates:
        if cand.lower().strip() in cols_map:
            return cols_map[cand.lower().strip()]
    for col_lower, orig in cols_map.items():
        for cand in candidates:
            if cand.lower().strip() in col_lower:
                return orig
    return None

def try_parse_dates(series, dayfirst=True):
    parsed = pd.to_datetime(series, dayfirst=dayfirst, errors="coerce")
    if parsed.isna().mean() > 0.2:
        alt = pd.to_datetime(series, dayfirst=not dayfirst, errors="coerce")
        if alt.isna().mean() < parsed.isna().mean():
            parsed = alt
    return parsed

def clean_numeric(x):
    if pd.isna(x): return pd.NA
    if isinstance(x, (int, float)): return x
    s = re.sub(r"[^\d\.\-]", "", str(x).strip())
    try: return float(s)
    except: return pd.NA

def safe_index(columns, value):
    if value is None: return 0
    try: return list(columns).index(value) + 1
    except: return 0

def pretty_date(k):
    try: return datetime.strptime(str(k), "%Y-%m-%d").strftime("%d-%m-%Y")
    except: return str(k)


# ── EXCEL BUILDER WITH MERGED DATE HEADERS ───────────────────────────────────

def to_excel_bytes_multilevel(emp_col_name, name_col_name, deduped, date_cols_sorted):
    """
    Build Excel with two header rows:
    Row 1: EMP ID | Employee Name | [DATE merged across 3 cols] | [DATE merged across 3 cols] ...
    Row 2:                         | Total CTC | Fixed CTC | Total Variable Pay | ...
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "CT Pivoted"

    # ── Colour palette ───────────────────────────────────────────────────────
    DARK       = "1A1714"   # hero background
    ORANGE     = "D4521A"   # brand accent
    LIGHT_HDR  = "F5F0EB"   # text on dark
    DATE_ALT   = "FDF3EE"   # alternating date group bg (light orange tint)
    DATE_ALT2  = "FFFFFF"   # alternating date group bg (white)
    SUBHDR_BG  = "3D3530"   # sub-header row bg (dark brown)
    SUBHDR_FG  = "F5F0EB"   # sub-header row text

    thin  = Side(style="thin",   color="D4C8BC")
    thick = Side(style="medium", color=ORANGE)
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_font(bold=True, color=LIGHT_HDR, size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def cell_font(bold=False, size=10):
        return Font(name="Calibri", bold=bold, size=size)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    # ── Build data matrix ────────────────────────────────────────────────────
    # Pivot: index=(EMP_ID, EMP_NAME), columns=WHEN_KEY, values={TOTAL_CT, FIXED_CT, VAR_PAY}
    deduped_s = deduped.sort_values(["EMP_ID", "WHEN_KEY"])

    pivot_total = deduped_s.pivot(index=["EMP_ID", "EMP_NAME"], columns="WHEN_KEY", values="TOTAL_CT_NUM").reset_index()
    pivot_fixed = deduped_s.pivot(index=["EMP_ID", "EMP_NAME"], columns="WHEN_KEY", values="FIXED_CT_NUM").reset_index()
    pivot_var   = deduped_s.pivot(index=["EMP_ID", "EMP_NAME"], columns="WHEN_KEY", values="VAR_PAY_NUM").reset_index()

    pivot_total.columns.name = None
    pivot_fixed.columns.name = None
    pivot_var.columns.name   = None

    # Align all three pivots to same employee list
    emp_base = pivot_total[["EMP_ID", "EMP_NAME"]].copy()

    # ── ROW 1: top headers ────────────────────────────────────────────────────
    # Col 1 = EMP ID, Col 2 = Employee Name, then 3 cols per date
    n_dates   = len(date_cols_sorted)
    total_cols = 2 + n_dates * 3

    # Write Row 1
    ws.cell(1, 1, emp_col_name).font  = header_font()
    ws.cell(1, 1).fill                = fill(DARK)
    ws.cell(1, 1).alignment           = center_align
    ws.cell(1, 1).border              = border_thin

    ws.cell(1, 2, name_col_name).font = header_font()
    ws.cell(1, 2).fill                = fill(DARK)
    ws.cell(1, 2).alignment           = center_align
    ws.cell(1, 2).border              = border_thin

    # Merge rows 1-2 for EMP ID and Employee Name columns
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    for i, dk in enumerate(date_cols_sorted):
        date_label = pretty_date(dk)
        col_start  = 3 + i * 3
        col_end    = col_start + 2
        # Alternating bg for date groups
        bg = DATE_ALT if i % 2 == 0 else DATE_ALT2
        fg = ORANGE if i % 2 == 0 else DARK

        ws.cell(1, col_start, date_label).font      = Font(name="Calibri", bold=True, color=fg, size=10)
        ws.cell(1, col_start).fill                  = fill(bg)
        ws.cell(1, col_start).alignment             = center_align
        ws.cell(1, col_start).border                = Border(left=thick, right=thick, top=thick, bottom=thin)
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)

    # ── ROW 2: sub-headers ────────────────────────────────────────────────────
    sub_labels = ["Total CTC", "Fixed CTC", "Total Variable Pay"]
    for i, dk in enumerate(date_cols_sorted):
        col_start = 3 + i * 3
        bg = DATE_ALT if i % 2 == 0 else DATE_ALT2
        for j, lbl in enumerate(sub_labels):
            c = ws.cell(2, col_start + j, lbl)
            c.font      = Font(name="Calibri", bold=True, color=DARK, size=9)
            c.fill      = fill("E8DDD6" if i % 2 == 0 else "F0EDED")
            c.alignment = center_align
            left_b  = thick if j == 0 else thin
            right_b = thick if j == 2 else thin
            c.border = Border(left=left_b, right=right_b, top=thin, bottom=thick)

    # ── DATA ROWS ─────────────────────────────────────────────────────────────
    for row_idx, (_, emp_row) in enumerate(emp_base.iterrows()):
        r    = row_idx + 3  # Excel row (1-indexed, rows 1&2 are headers)
        eid  = emp_row["EMP_ID"]
        ename = emp_row["EMP_NAME"]
        row_bg = "FFFFFF" if row_idx % 2 == 0 else "FAF8F6"

        c1 = ws.cell(r, 1, eid)
        c1.font      = cell_font(bold=True)
        c1.fill      = fill(row_bg)
        c1.alignment = left_align
        c1.border    = border_thin

        c2 = ws.cell(r, 2, ename)
        c2.font      = cell_font()
        c2.fill      = fill(row_bg)
        c2.alignment = left_align
        c2.border    = border_thin

        for i, dk in enumerate(date_cols_sorted):
            col_start = 3 + i * 3
            bg = "FDF8F6" if i % 2 == 0 else row_bg

            def get_val(pivot_df, key, emp_id):
                row = pivot_df[pivot_df["EMP_ID"] == emp_id]
                if row.empty or key not in row.columns:
                    return ""
                v = row.iloc[0][key]
                if pd.isna(v): return ""
                try: return int(v) if float(v) == int(float(v)) else float(v)
                except: return v

            vals = [
                get_val(pivot_total, dk, eid),
                get_val(pivot_fixed, dk, eid),
                get_val(pivot_var,   dk, eid),
            ]
            for j, val in enumerate(vals):
                c = ws.cell(r, col_start + j, val)
                c.font      = cell_font()
                c.fill      = fill(bg)
                c.alignment = right_align if isinstance(val, (int, float)) and val != "" else left_align
                if isinstance(val, (int, float)) and val != "":
                    c.number_format = '#,##0'
                left_b  = thick if j == 0 else thin
                right_b = thick if j == 2 else thin
                c.border = Border(left=left_b, right=right_b, top=thin, bottom=thin)

    # ── COLUMN WIDTHS ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    for i in range(n_dates * 3):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    # ── FREEZE PANES ──────────────────────────────────────────────────────────
    ws.freeze_panes = "C3"

    # ── ROW HEIGHTS ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HERO ──────────────────────────────────────────────────────────────────────

admin_badge = '<span class="admin-badge">Admin</span>' if is_admin else ""
st.markdown(f"""
<div class="hero">
  <div class="hero-user">{'👑 ' if is_admin else ''}<strong>{user_email}</strong>{admin_badge}</div>
  <div class="hero-label">Payroll Internal Tool</div>
  <div class="hero-title">CT Data <strong>Transformer</strong></div>
  <div class="hero-sub">Upload your employee CT file. Keeps the latest record per change date
  and pivots — one row per employee, each change date as a grouped column (Total CTC · Fixed CTC · Variable Pay).</div>
</div>
""", unsafe_allow_html=True)

_, signout_col = st.columns([6, 1])
with signout_col:
    if st.button("Sign Out", key="logout"):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

if is_admin:
    with st.expander("⚙️ Admin — Manage Users", expanded=False):
        st.markdown("**Add/remove users in Streamlit Secrets** → [users] section:")
        st.code("""[users]
"vivek.naiwal@cars24.com" = "Vivek@007"
"payroll1@cars24.com"     = "Password1"
"payroll2@cars24.com"     = "Password2"
"manager@cars24.com"      = "Password3"
""", language="toml")
        st.info("Go to Streamlit Cloud → your app → ⋮ → Settings → Secrets to add/remove users.")
        st.markdown("**Currently signed-in user:** " + user_email)

st.divider()

# ── STEP 1: UPLOAD ────────────────────────────────────────────────────────────

st.markdown('<div class="section-head"><span>01</span> Upload File</div>', unsafe_allow_html=True)

uploaded = st.file_uploader("Drop your CSV or Excel file here",
    type=["csv", "xlsx", "xls"], accept_multiple_files=False, label_visibility="collapsed")

if uploaded is None:
    st.info("📂 Upload a file above to get started.")
    st.stop()

try:
    raw_bytes = uploaded.read()
    if uploaded.name.lower().endswith(".csv"):
        df = pd.read_csv(io.BytesIO(raw_bytes), dtype=str)
    else:
        df = pd.read_excel(io.BytesIO(raw_bytes), dtype=str)
    df.columns = [c.strip() for c in df.columns]
except Exception as e:
    st.error(f"❌ Failed to read file: {e}")
    st.stop()

st.success(f"✅ {uploaded.name} loaded — {len(df):,} rows · {len(df.columns)} columns")
st.divider()

# ── STEP 2: CONFIGURE ─────────────────────────────────────────────────────────

st.markdown('<div class="section-head"><span>02</span> Configure</div>', unsafe_allow_html=True)

col_opt1, col_opt2, col_opt3 = st.columns(3)
with col_opt1: dayfirst   = st.checkbox("Dates are DD-MM-YYYY (day first)", value=True)
with col_opt2: auto_clean = st.checkbox("Clean numeric fields (remove commas/currency)", value=True)
with col_opt3: drop_empty = st.checkbox("Drop all-empty date columns in output", value=True)

fc1, fc2 = st.columns([1, 3])
with fc1:
    filter_date = st.date_input("Remove Change Dates before",
        value=pd.Timestamp("2024-01-01"),
        min_value=pd.Timestamp("2000-01-01"),
        max_value=pd.Timestamp("2030-12-31"),
        help="Rows with Change Date before this are excluded")

detected = {k: find_column(df.columns.tolist(), v) for k, v in COLUMN_CANDIDATES.items()}
options_with_none = [None] + list(df.columns)

st.markdown("**Column Mapping** — auto-detected from your headers:")
mc1, mc2, mc3, mc4, mc5, mc6, mc7 = st.columns(7)
with mc1: emp_col      = st.selectbox("EMP ID",            options_with_none, index=safe_index(df.columns, detected["emp_id"]))
with mc2: name_col     = st.selectbox("Employee Name",     options_with_none, index=safe_index(df.columns, detected["emp_name"]))
with mc3: when_col     = st.selectbox("When Was Change?",  options_with_none, index=safe_index(df.columns, detected["when_change"]))
with mc4: total_col    = st.selectbox("Total CTC",         options_with_none, index=safe_index(df.columns, detected["total_ctc"]))
with mc5: fixed_col    = st.selectbox("Fixed CTC",         options_with_none, index=safe_index(df.columns, detected["fixed_ctc"]))
with mc6: var_col      = st.selectbox("Total Variable Pay",options_with_none, index=safe_index(df.columns, detected["variable_pay"]))
with mc7: created_col  = st.selectbox("Created Date",      options_with_none, index=safe_index(df.columns, detected["created_date"]))

missing = [k for k, v in {
    "EMP ID": emp_col, "Employee Name": name_col,
    "When Was Change?": when_col, "Total CTC": total_col,
    "Fixed CTC": fixed_col, "Total Variable Pay": var_col,
    "Created Date": created_col
}.items() if not v]

if missing:
    st.error(f"❌ Please map these columns: {', '.join(missing)}")
    st.stop()

st.divider()

# ── STEP 3: PROCESS ───────────────────────────────────────────────────────────

st.markdown('<div class="section-head"><span>03</span> Process</div>', unsafe_allow_html=True)

if st.button("⚡  Process & Generate Output", type="primary", use_container_width=True):
    with st.spinner("Processing..."):

        work = df[[emp_col, name_col, when_col, total_col, fixed_col, var_col, created_col]].copy()
        work.columns = ["EMP_ID", "EMP_NAME", "WHEN_CHG", "TOTAL_CT", "FIXED_CT", "VAR_PAY", "CREATED"]

        work["WHEN_DT"]    = try_parse_dates(work["WHEN_CHG"], dayfirst=dayfirst)
        work["CREATED_DT"] = try_parse_dates(work["CREATED"],  dayfirst=dayfirst)

        cutoff = pd.Timestamp(filter_date)
        before = work["WHEN_DT"] < cutoff
        dropped = before.sum()
        work = work[~before].copy()
        if dropped > 0:
            st.info(f"🗑️ Removed {dropped:,} rows before {cutoff.strftime('%d-%m-%Y')} · {len(work):,} rows remaining")

        work["WHEN_KEY"] = work["WHEN_DT"].dt.strftime("%Y-%m-%d")
        bad_mask = work["WHEN_KEY"].isna()
        work.loc[bad_mask, "WHEN_KEY"] = work.loc[bad_mask, "WHEN_CHG"].astype(str).str.strip()

        work["CRT_INT"] = work["CREATED_DT"].values.astype("datetime64[ns]").astype("int64")
        work.loc[work["CREATED_DT"].isna(), "CRT_INT"] = -9223372036854775808

        if auto_clean:
            work["TOTAL_CT_NUM"] = work["TOTAL_CT"].apply(clean_numeric)
            work["FIXED_CT_NUM"] = work["FIXED_CT"].apply(clean_numeric)
            work["VAR_PAY_NUM"]  = work["VAR_PAY"].apply(clean_numeric)
        else:
            work["TOTAL_CT_NUM"] = pd.to_numeric(work["TOTAL_CT"], errors="coerce")
            work["FIXED_CT_NUM"] = pd.to_numeric(work["FIXED_CT"], errors="coerce")
            work["VAR_PAY_NUM"]  = pd.to_numeric(work["VAR_PAY"],  errors="coerce")

        # Dedup: keep latest Created Date per (EMP, change_date)
        idx     = work.groupby(["EMP_ID", "WHEN_KEY"])["CRT_INT"].idxmax()
        deduped = work.loc[idx].copy()
        deduped = deduped.sort_values(["EMP_ID", "WHEN_KEY"])

        # Sorted unique date keys
        all_date_keys = sorted(deduped["WHEN_KEY"].dropna().unique())

        if drop_empty:
            # Only keep dates where at least one employee has a non-null Total CTC
            all_date_keys = [
                dk for dk in all_date_keys
                if deduped[deduped["WHEN_KEY"] == dk]["TOTAL_CT_NUM"].notna().any()
            ]

        # Build preview dataframe (flat) for on-screen display
        preview_rows = []
        for eid, grp in deduped.groupby("EMP_ID"):
            row = {emp_col: eid, name_col: grp["EMP_NAME"].iloc[0]}
            for dk in all_date_keys:
                match = grp[grp["WHEN_KEY"] == dk]
                dl    = pretty_date(dk)
                if not match.empty:
                    row[f"{dl} · Total CTC"]  = match.iloc[0]["TOTAL_CT_NUM"]
                    row[f"{dl} · Fixed CTC"]  = match.iloc[0]["FIXED_CT_NUM"]
                    row[f"{dl} · Var Pay"]    = match.iloc[0]["VAR_PAY_NUM"]
                else:
                    row[f"{dl} · Total CTC"]  = None
                    row[f"{dl} · Fixed CTC"]  = None
                    row[f"{dl} · Var Pay"]    = None
            preview_rows.append(row)
        preview_df = pd.DataFrame(preview_rows).fillna("")

        xlsx_bytes = to_excel_bytes_multilevel(emp_col, name_col, deduped, all_date_keys)

        st.session_state["xlsx_bytes"]    = xlsx_bytes
        st.session_state["preview_df"]    = preview_df
        st.session_state["deduped"]       = deduped
        st.session_state["total"]         = len(df)
        st.session_state["n_dates"]       = len(all_date_keys)

# ── RESULTS ───────────────────────────────────────────────────────────────────

if "preview_df" in st.session_state:
    preview_df = st.session_state["preview_df"]
    deduped    = st.session_state["deduped"]
    total      = st.session_state["total"]
    n_dates    = st.session_state["n_dates"]

    n_emp   = len(preview_df)
    n_dedup = len(deduped)

    st.divider()
    st.markdown(f"""
    <div class="metric-row">
      <div class="metric-card"><div class="metric-val">{total:,}</div><div class="metric-label">Input Rows</div></div>
      <div class="metric-card"><div class="metric-val">{n_dedup:,}</div><div class="metric-label">After Dedup</div></div>
      <div class="metric-card"><div class="metric-val">{n_emp:,}</div><div class="metric-label">Employees</div></div>
      <div class="metric-card"><div class="metric-val">{n_dates}</div><div class="metric-label">Date Groups</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.success(f"✅ Done! {n_emp:,} employees · {n_dates} change-date groups · 3 CTC fields per date")

    st.markdown('<div class="section-head"><span>04</span> Preview (first 15 rows — flat view)</div>', unsafe_allow_html=True)
    st.dataframe(preview_df.head(15), use_container_width=True, hide_index=True)
    st.caption("ℹ️ Preview shows flat column names. The downloaded Excel has proper merged date headers with 3 sub-columns per date.")

    st.markdown('<div class="section-head"><span>05</span> Download</div>', unsafe_allow_html=True)
    st.download_button(
        "⬇ Download Excel (.xlsx) — with merged date headers",
        data=st.session_state["xlsx_bytes"],
        file_name="employee_ct_pivoted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_xlsx"
    )

    with st.expander("🔍 Diagnostics"):
        st.dataframe(
            deduped[["EMP_ID", "EMP_NAME", "WHEN_CHG", "CREATED", "TOTAL_CT_NUM", "FIXED_CT_NUM", "VAR_PAY_NUM"]].head(20),
            use_container_width=True, hide_index=True
        )
        dup = deduped.groupby(["EMP_ID", "WHEN_KEY"]).size().reset_index(name="count")
        dup = dup[dup["count"] > 1].sort_values("count", ascending=False)
        st.write("Duplicate keys (top 10):" if len(dup) else "✅ No duplicates found.")
        if len(dup):
            st.dataframe(dup.head(10), use_container_width=True, hide_index=True)
